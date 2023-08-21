from pykrx import stock, bond
from datetime import datetime, timedelta
import time
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import ExDriver as ed   #엑셀 파일 입출력 전용함수.

pathTok = "\\"
class StockKr:
    def __init__(self):
        self.now = datetime.now()
        self.today = "{0:0>2}{1:0>2}{2:0>2}".format(self.now.year, self.now.month, self.now.day)
        self.cwd = os.getcwd()

        # 주식코드 : 주식명 --> local file로 저장
        self.tk_KOSPI_tkdict = dict()
        self.tk_KOSDAQ_tkdict = dict()

        # 데이터를 읽어오는 딕셔너리.
        self.thema_KOSPI_tkdict = dict()      #테마주(관심주) 중 코스피 데이터 가져오기.
        self.thema_KOSDAQ_tkdict = dict()     #테마주(관심주) 중 코스닥 데이터 가져오기.

        self.thema_total_dict = dict()        #테마주(관심주) 중 코스피 + 코스닥 데이터 가져오기.
        self.force_total_dict = dict()        #오늘의 세력주(거래량폭등주) 데이터 가져오기. <-- 기본 상위 10개

        self.subfile_path = os.path.join(os.getcwd(), "서브파일")
        self.data_path = {
            "주식코드코스피" : "StockCode_KOSPI.txt",
            "주식코드코스닥" : "StockCode_KOSDAQ.txt",
            "관심주" : "WantCode.txt",

            "일봉": "dayinfo.xlsx",
            "일봉거래공매" : "dayinfosub.xlsx",
            
            "세력주": "SmallCompony.xlsx",
            "세력주_일봉": "SC_dayinfo.xlsx"
        }
    def module (self, code_update=False, dayinfo_update=True, dayinfo_sub_update=True):
        if not os.path.isdir(self.subfile_path):
            os.mkdir(self.subfile_path)
        self.module_readTr(update=code_update)
        if dayinfo_update == True:
            print("\n\n 일봉데이터를 조회합니다 \n\n")
            self.update_day_chart()
        if dayinfo_sub_update == True:
            print("\n\n 공매도, 거래량 데이터를 조회합니다 \n\n")
            self.update_day_chart(sub=True)

    def get_stock_manager (self, start : str, end : str, tr_code : str, mode="d", sub=False):
        ''' 원래뽑는 종가 + 시가총액 등 추가 정보 get
        :param start:
        :param end:
        :param tr_code:
        :param mode:
        :param sub : (공매도, 기관) 데이터를 가져오는지 확인.
        :return:
        '''
        if sub == True:
            print("{공매도 정보를 가져오는중 ...}")
            df = stock.get_shorting_balance_by_date(start, end, tr_code)
            return df
        else:
            print("{종가,시가,고가,거래량을 가져오는 중 ...}")
            df1 = stock.get_market_ohlcv(start, end, tr_code, mode)
            print("{시가총액 정보를 가져오는 중 ...}")
            df2 = stock.get_market_cap(start, end, tr_code, mode)
            print("{매수 거래량을 가져오는 중 ...}")
            df3 = stock.get_market_trading_volume_by_date(start, end, tr_code)
            name_list = []
            for item in df3.columns:
                name_list.append(str(item) + "_매수")
            df3.columns = name_list
            print("{매도 거래량을 가져오는 중 ...}")
            df4 = stock.get_market_trading_volume_by_date(start, end, tr_code, on="매도")
            name_list2 = []
            for item in df4.columns:
                name_list2.append(str(item) + "_매도")
            df4.columns = name_list2
            df = pd.concat([df1, df2, df3, df4], axis=1)
            df = df.loc[:, ~df.T.duplicated()]
            if len(df) != 0:
                print(df.tail(5))
            return df

    def init_get_day_info (self, tr : list, path : str, sub : bool):
        ''' 주식데이터가 하나도 없을 시 사용하는 매소드 (60일씩 4번 --> 240일전부터 데이터가져옴.)
        :param tr: [Company, Tr_code]
        :param path : 파일경로
        :param sub : False(일봉정보), True(거래량공매도정보)
        :return:
        '''
        #self.df_to_file(df, self.cwd + pathTok + self.data_path["일봉"], company)
        base = self.day_counter(start=self.today, offset=120)
        while base <= self.today:
            df = self.get_day_stockinfo(tr=tr, start=base, offset=80, pos=1, sub=sub)
            if len(df) == 0:
                base = self.day_counter(start=base, offset=80, pos=1)
                print("Not data")
                continue
            base = self.day_counter(self.df_date_to_str(df), offset=2, pos=1)
            self.df_to_file(df, path=path, company=tr[0])

    def update_day_chart(self, sub=False):
        day_chart_path = os.path.join(self.subfile_path, self.data_path["일봉"])
        if sub == True:
            day_chart_path = os.path.join(self.subfile_path, self.data_path["일봉거래공매"])
        #파일이 없는 경우 --> 초기화세팅
        if not os.path.isfile(day_chart_path):
            for company, tr_code in self.thema_total_dict.items():
                #self.get_day_stockinfo(tr=[company, tr_code], start=self.today, offset=120, pos=-1)
                self.init_get_day_info(tr=[company, tr_code], path = day_chart_path, sub=sub)
        else:
            for company, tr_code in self.thema_total_dict.items():
                df = self.read_dayinfo(day_chart_path, company)
                if len(df) == 0:
                    #self.get_day_stockinfo(tr=[company, tr_code], start=self.today, offset=120, pos=-1)
                    self.init_get_day_info(tr=[company, tr_code], path = day_chart_path, sub=sub)
                else:
                    saved_last_day =self.df_date_to_str(df, offset=1, whence="tail")
                    saved_last_day_pp = self.day_counter(start=saved_last_day, offset=2, pos=1)
                    if self.today < saved_last_day:
                        continue

                    print("[\'" + company + "\' 현재 기준일(" + saved_last_day_pp + ")로부터 ("
                                + self.today + ")일 까지의 일 별 데이터를 가져오는중...]")
                    df = self.get_stock_manager(start = saved_last_day_pp, end = self.today, tr_code=tr_code, sub=sub)
                    if len(df) == 0 or self.df_date_to_str(df) <= saved_last_day:
                        continue
                    self.df_to_file(df, day_chart_path, company)

    def df_date_to_str(self, df: pd.DataFrame, whence="tail", offset=1):
        ''' Data Frame의 날짜를 정수형으로 반환해주는 함수.
        :param df:
        :param whence: tail(끝에서부터), head(앞에서부터)
        :param offset: 몇번째 index?
        :return:
        '''
        if whence == "tail":
            return str(df.tail(offset).index.values[0])[:10].replace("-", "")  # 마지막 행 value 값 구함.
        elif whence == "head":
            return str(df.head(offset).index.values[0])[:10].replace("-", "")  # 마지막 행 value 값 구함.
        else:
            return -1

    def read_dayinfo(self, path: str, company: str):
        ''' 회사별 일봉 데이터차트를 Dataframe
        :param path: 경로
        :param company: reading할 회사
        :return: 일봉 Dataframe
        '''
        #회사명이 없는 경우.. -> -1 리턴
        finding_sheet = openpyxl.load_workbook(filename=path)
        com_list = finding_sheet.sheetnames
        if not company in com_list:
            return pd.DataFrame()

        df_sheet_idx = pd.read_excel(path, sheet_name=company, engine="openpyxl")
        df_sheet_idx.set_index("날짜", inplace=True)
        return df_sheet_idx

    def df_to_file(self, df : pd.DataFrame, path : str, company : str):
        ''' 데이터프레임(df) 를 path 경로에 저장
        :param df: 데이터프레임
        :param df_sheet_idx.tail(1): 경로.
        :param company : 엑셀시트명 (회사명)
        :return:
        '''

        global wb
        if not os.path.isfile(path):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(filename=path)
        wb_list = wb.sheetnames
        print("{Sheet Extract : \"" + company +"\" sheet Data Frame ... }")
        df.reset_index(inplace=True)           # 파이썬에서 날짜 인덱스만 따로 빠져서 저장되는 현상 때문에 해당 인덱스를 원상태로 복구
        if "Sheet" in wb_list:
            del wb["Sheet"]        #삭제 시에는 엑셀 객체 wb[] 로 인덱스 번호로 삭제해야함
        if not company in wb_list:
            wb.create_sheet(company)
            ws = wb[company]
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        else:
            ws = wb[company]
            for r in dataframe_to_rows(df, index=False, header=False):   # Header는 제목
                ws.append(r)
        try:
            wb.save(path)
        except PermissionError:
            print("[파일 저장 오류!] 열고있는 엑셀파일을 닫아주세요!!!!!")
            time.sleep(5)
            self.df_to_file (df, path, company) #닫길때까지 재귀호출

    def day_counter(self, start="default", offset=60, pos=-1):
        ''' 날짜계산기 : start 기준으로 몇 일 offset만큼 갈 것인가?
        :param start:
        :param offset: 가고싶은 날짜, ex) start=20230729, offset=4, pos=-1 -> 20230726
        :param pos: 1 (증가하는방향), -1 (감소하는방향)
        :return: 계산된 날짜 문자열
        '''
        if start == "default":
            start = self.today
        cur_year = start[:4]
        cur_mon = start[4:6]
        cur_day = start[6:]

        s_time = datetime(year=int(cur_year), month=int(cur_mon), day=int(cur_day))
        off = (offset-1) * pos
        s_time += timedelta(days=off)
        ret_addr = "{0:0>2}{1:0>2}{2:0>2}".format(s_time.year, s_time.month, s_time.day)
        return ret_addr
    def day_counter_offset(self, start, end):
        ''' start와 end 의 날짜 계산'''
        s_y, s_m, s_d = start[:4], start[4:6], start[6:]
        e_y, e_m, e_d = end[:4], end[4:6], end[6:]
        time_offset = datetime(year=int(e_y), month=int(e_m), day=int(e_d)) -\
                      datetime(year=int(s_y), month=int(s_m), day=int(s_d)) + timedelta(days=1)
        return time_offset.days
    def get_day_stockinfo(self, tr : list, start="default", offset=90, pos=-1, sub=False):
        ''' 하루별 want data에 대하여 date_list만큼 주식 데이터를 가져옴
        :param tr : [회사명, Tr_code]
        :param start: 시작 날짜
        :param offset: offset: 가고싶은 날짜, ex) start=20230729, offset=4, pos=-1 -> 20230726
        :param pos: pos: 1 (증가하는방향), -1 (감소하는방향)
        :param sub : False(일봉정보), True(공매,거래정보)
        :return: 가져온 df
        '''
        company, s_code = tr
        if start == "default":
            start = self.today
        print("[일별 주식 데이터를 가져오는중...]")
        end = self.day_counter(start, offset, pos)
        if pos == -1:
            tmp = start
            start = end
            end = tmp

        print("[\'" + company + "\' 현재 기준일(" + start + ")부터 " + str(offset) + "간격의 일 별 데이터를 가져오는중...]")
        df = self.get_stock_manager(start=start, end=end, tr_code=s_code, sub=sub)
        return df

    ''' 주식종목 코드 및 원하는 주식 종목 추출 메소드류'''
    def module_readTr(self, update=False):
        ''' 현재 보고싶은 종목(회사명) --> 종목코드로 전환 및, 종목코드 갱신/저장에 관한 모듈
        :param update: False(갱신안함), True(갱신함)
        :return:
        '''
        path_code_PI = os.path.join(self.subfile_path, self.data_path["주식코드코스피"])
        path_code_DAQ = os.path.join(self.subfile_path, self.data_path["주식코드코스닥"])
        if update == True or not os.path.isfile(path_code_PI) or os.path.getsize(path_code_DAQ) == 0:
            self.extract_ticker_localfile(code_path=path_code_PI, market="KOSPI")
        if update == True or not os.path.isfile(path_code_DAQ) or os.path.getsize(path_code_DAQ) == 0:
            self.extract_ticker_localfile(code_path=path_code_DAQ, market="KOSDAQ")
        self.read_ticker_localfile()
        self.read_tickerThema_localfile()

    def str_lower_space (self, strs : str):
        """ strs 를 공백문자없고 + 소문자화하여 저장
        :param strs: 회사명(가공전)
        :return:
        """
        strs = strs.replace(" ", "").lower()
        return strs

    def extract_ticker_localfile(self, code_path="StockCode.txt", market="ALL"):
        ''' 오늘 일자 기준으로 상장된 주식코드와 회사명을 .txt 파일로저장
        :param code_path: 종목코드를 저장할 위치
        :param market: KOSDAQ, KOSPI, ALL, ..
        :return:
        '''
        tk = stock.get_market_ticker_list(self.today, market)
        fp = open(code_path, "w", encoding="UTF-8")
        line = []
        print("상장 코드 갱신... : " + code_path)
        line.append(self.today)
        for t in tk:
            name = stock.get_market_ticker_name(t)
            line_str = str(name) + ":" + str(t)
            line.append(line_str)

        fp.writelines("\n".join(line))
        fp.close()
    def read_ticker_localfile(self):
        code_path_PI = os.path.join(self.subfile_path, self.data_path["주식코드코스피"])
        code_path_DAQ = os.path.join(self.subfile_path, self.data_path["주식코드코스닥"])


        print("[주식코드 가져오는중...] : " + code_path_PI + ", " + code_path_DAQ)
        fp_pi = open(code_path_PI, "r", encoding="UTF-8")
        fp_daq = open(code_path_DAQ, "r", encoding="UTF-8")
        lines_pi = fp_pi.readlines()
        fp_pi.close()
        lines_daq = fp_daq.readlines()
        fp_daq.close()

        if len(lines_pi) < 1 or len(lines_daq) < 1:
            return False

        update_date_pi = lines_pi[0][:-1]
        update_date_daq = lines_daq[0][:-1]
        print("[Update Date] : (KOSPI : " + update_date_pi + "), (KOSDAQ : " + update_date_daq + ")")
        for item in lines_pi[1:]:
            item = item[:-1].split(":")
            self.tk_KOSPI_tkdict[self.str_lower_space(item[0].strip())] = item[1].strip()

        for item in lines_daq[1:]:
            item = item[:-1].split(":")
            self.tk_KOSDAQ_tkdict[self.str_lower_space(item[0].strip())] = item[1].strip()

    def read_tickerThema_localfile (self):
        self.read_ticker_localfile()
        code_path = self.cwd + pathTok + self.data_path["관심주"]
        print("[테마주 가져오는 중...] : " + code_path)
        fp = open(code_path, "r", encoding="UTF-8")
        line = fp.readlines()
        fp.close()
        for l in line:
            l = l.replace("\n", "").strip()
            if "-" in l or len(l) == 0:
                continue
            l = self.str_lower_space(l)

            if l in self.tk_KOSPI_tkdict:
                self.thema_KOSPI_tkdict[l] = self.tk_KOSPI_tkdict[l]
            elif l in self.tk_KOSDAQ_tkdict:
                self.thema_KOSDAQ_tkdict[l] = self.tk_KOSDAQ_tkdict[l]
            else:
                print("[name error] : \'" + l + "\' 사명에 해당하는 주식코드가 없습니다.")
                continue

        print("KOSPI : ", end="")
        print(self.thema_KOSPI_tkdict)
        print("KOSDAQ : ", end="")
        print(self.thema_KOSDAQ_tkdict)

        # 전체 파일리스트 관리.
        self.thema_total_dict.update(self.thema_KOSPI_tkdict)
        self.thema_total_dict.update(self.thema_KOSDAQ_tkdict)



    def find_small_module(self, rank=10):
        ''' 세력주 (거래량급등주) 찾기 알고리즘 '''
        self.find_small_init(rank=rank)
        self.find_small_getDayinfo()

    def find_small_init(self, rank=10):
        today = self.today
        df = stock.get_market_cap(today)
        filename = os.path.join(self.subfile_path, self.data_path["세력주"])
        cost_dict = {}
        tmp_list = []
        low_cost = 200000000000
        high_cost = 500000000000

        for ticker in df.index:
            tr = df.loc[ticker, "시가총액"]
            if tr >= low_cost and tr <= high_cost:
                tmp_list.append([ticker, df.loc[ticker, "거래량"]])

        print("{거래량 정렬중..}")
        sorted_list = sorted(tmp_list, key=lambda x: -x[1])

        ticker_list = []
        print("{오늘의 세력주... } : " + self.today)
        #상위 10개만 골라오기
        for ticker, tr_size in sorted_list[:rank]:
            name = stock.get_market_ticker_name(ticker)
            print(name, tr_size)
            ticker_list.append([name, ticker, tr_size])
            self.force_total_dict[name] = ticker


        df = pd.DataFrame(data=ticker_list, columns=["회사명", "티커", "거래량"])
        ed.e_write(filename=filename, sname=today, df=df, last=False, header=True, index=False)

    def find_small_getDayinfo(self):
        filename = os.path.join(self.subfile_path, self.data_path["세력주_일봉"])
        exl = ed.SavedExcel(filename=filename)

        for compony, tr_code in self.force_total_dict.items():
            last_val = exl.read(last=True, sname=compony)
            if last_val == None:
                print("{"+ compony + "} 데이터가 없어 갱신하는중...")
            else:
                print("{" + compony + "} 데이터가 존재! 업데이트 중...")
                latest_date = last_val.index


        #df = ed.e_read(filename=filename, sname=)
if __name__ == "__main__":
    aa = StockKr()
    aa.find_small_module()
    #aa.module(code_update=True, dayinfo_update=True)