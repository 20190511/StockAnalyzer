import os
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import krxdata as krx
import talib as ta

def dt(year=0, mon=0, day=0, strs=""):
    ''' (2023,08,11) or "20230811" 을 datatime 객체로 Translate 함수.
    :param year:
    :param mon:
    :param day:

    :param str:
    :return:
    '''
    if strs != "":
        return datetime(year=int(strs[:4]), month=int(strs[4:6]), day=int(strs[6:]))
    else:
        return datetime(year=year, month=mon, day=day)
def df_t(df : pd.DataFrame, index_num : int):
    ''' DataFrame 인덱스값 범위를 계산해서 구해주는 함수.
    '''
    if index_num < 0 or index_num >= len(df):
        return -1
    return df.loc[index_num]
def writeExcelFromDf(df: pd.DataFrame, path: str, sname: str, index=False):
    if len(df) == 0:
        return -1
    if not os.path.isfile(path):
        wb = openpyxl.Workbook()
        wb.create_sheet(sname)
        ws = wb[sname]
        for r in dataframe_to_rows(df, index=index, header=True):
            ws.append(r)
        wb.save(path)
    else:
        wb = openpyxl.load_workbook(filename=path)
        wb_list = wb.sheetnames
        if "Sheet" in wb_list:
            del wb["Sheet"]

        header=False
        if not sname in wb_list:
            wb.create_sheet(sname)
            header = True
        ws = wb[sname]
        for r in dataframe_to_rows(df, index=index, header=header):
            ws.append(r)
        wb.save(path)
def readExcelToDf(path: str, sname: str):
    ''' 액샐 파일에서 데이터 추출
    :param path : 파일 이름
    :param sname : 해당 엑셀 파일 시트 이름
    :return: 엑셀에서 추출한 DataFrame (*없으면 len(df) == 0)
    '''
    if not os.path.isfile(path):
        return pd.DataFrame()  # NULL 데이터프레임 리턴
    wb = openpyxl.load_workbook(filename=path)
    com_list = wb.sheetnames

    if not sname in com_list:
        return pd.DataFrame()

    df_sheet_idx = pd.read_excel(path, sheet_name=sname, engine="openpyxl")
    return df_sheet_idx
def add_excl_column(path: str,  df: pd.DataFrame, sheets :str, append=False):
    ''' 엑셀 열에 데이터 추가 데이터 추가 '''
    if not os.path.isfile(path):
        df.to_excel(path, sheet_name=sheets, index=False, header=True)
        return
    extract_df = df
    if append == True:
        origin_df = readExcelToDf(path=path, sname=sheets)
        extract_df = df_unify(origin_df, df)
    wb = openpyxl.load_workbook(filename=path)

    if not sheets in wb.sheetnames:
        wb.create_sheet(sheets)
    ws = wb[sheets]
    ws.delete_rows(1, ws.max_row)  # 기존 데이터 삭제
    for row in dataframe_to_rows(extract_df, index=False, header=True):
        ws.append(row)
    wb.save(path)
def df_check_row(df: pd.DataFrame, row_name: str):
    ''' DataFrame에 해당 row_name이 존재하는지 여부
    '''
    df_row_list = df.columns
    return row_name in df_row_list
def df_unify (*dfs):
    ''' DataFrame을 합쳐주는 함수.
        ex) df1, df2, df3 데이터를 df로 합쳐줌.'''
    df = pd.concat(list(dfs), axis=1)
    return df.loc[:, ~df.T.duplicated()]
#필요없으나 사용할 수도 있음
def df_slice(df:pd.DataFrame, data_col_list=["종가"], window=6, count=0):
    ''' DataFrame 을 window크기로 청크로 분할한 값을 리턴
    :param df:
    :param data_col_list: 청크에 포함 시킬 데이터 리스트
    :param window: 청크를 자를 단위 ex) 6 이면 6개 크기로..
    :param count: 몇 번째 인덱스 슬라이싱?
    :return:
    '''

    s, e = count, count + window
    #범위를 벗어나면 빈 데이터프레임 리턴.
    if s < 0 or e > len(df):
        return pd.DataFrame()

    ret_df = df_section(df=df[s:e], data_col_list=data_col_list)
    return ret_df
#필요없으나 사용할 수도 있음.
def df_section(df:pd.DataFrame, data_col_list=["종가"]):
    ''' DataFrame 중 data_col_list 리스트 내부의 섹션만 추출함.
    :param df:
    :param data_col_list:
    :return:
    '''
    ret_df = pd.DataFrame()
    ret_df["날짜"] = df["날짜"]
    for item in data_col_list:
        if df_check_row(df=df, row_name=item):
            ret_df[item] = df[item]
    return ret_df

def delete_empty_sheet(path: str):
    ''' 빈 시트 지우는 함수'''
    if "xlsx" in path and os.path.isfile(path):
        wb = openpyxl.load_workbook(filename=path)
        sheet_list = wb.sheetnames
        if "Sheet" in sheet_list:
            del wb["Sheet"]
        wb.save(path)


class StockAnaly:
    def __init__(self):
        #pd.set_option('display.max_columns', None)
        self.mykrx = krx.StockKr()

        ''' Path Info
            "주식코드코스피" : "StockCode_KOSPI.txt",
            "주식코드코스닥" : "StockCode_KOSDAQ.txt",
            "관심주" : "WantCode.txt",

            "일봉": "dayinfo.xlsx",
            "일봉거래공매" : "dayinfosub.xlsx"
        '''
        self.data_path = self.mykrx.data_path
        self.analy_path = {
            "지표": "StockCriteria.xlsx",
            "분석": "StockAnaly.xlsx",
            "점수": "StockScore.xlsx"
        }
        self.cwd = self.mykrx.cwd
        self.pathTok = krx.pathTok
        self.anal_namedict = {
            "SMA60_check": "60일 이동평균선 추이",
            "전기_nearess_check(후행)": "전환선 기준선 가까움(후행)",
            "전기_nearess_check": "전환선 기준선 가까움",
            "MACD_check": "MACD 상태",
            "후행스팬_line_cross_check": "후행스팬 x 전환선_기준선",
            "후행스팬_bong_cross_check": "후행스팬 x 봉",
            "스팬꼬리_check": "선행스팬 꼬리방향",
            "스팬위치_check": "봉과 구름대",
            "전_cross_기": "전환선 x 기준선",
            "봉_cross_전기": "봉 x 전환선_기준선",
        }
        self.anal_namedict_r = dict()           #역으로 구성된 딕셔너리.
        self.saved_df = pd.DataFrame()          #분석 or 지표계산할 때 총괄 데이터프레임

        self.anal_score = {"O": 0,
                 "up_near": 1,
                 "up_cross": 2,
                 "up": 3,
                 "down_near": 4,
                 "down_cross": 5,
                 "down": 6,
                 "X": 7,
                 "mid": 8}
        self.anal_scoreboard = dict()

        #이동평균선 리스트
        self.sma_window = [5,10,20,60,120]
        self.ema_window = [9,12,26]

        #최고가 관련
        self.high_crit = [9, 26]

        # 분석용 DataFrame
        self.read_df_dayinfo = pd.DataFrame()   # 일봉 데이터 가져오는 멤버변수
        self.read_df_criteria = pd.DataFrame()  # 지표 데이터 가져오는 멤버변수

        #점수 측정용 DataFrame
        self.today_score = dict()               # {key=점수 : Value=점수}
    def analy_dict_update(self):
        # X일 이동평균선 통과
        for item in self.sma_window:
            sma_name = "SMA" + str(item) + "_cross_check"
            dict_name = str(item) + "일 이동평균선 통과여부"
            self.anal_namedict[sma_name] = dict_name

        #x이 최고가 통과여부
        for item in self.high_crit:
            sma_name = str(item) + "_highest_check"
            dict_name = str(item) + "일 최고가 추이"
            self.anal_namedict[sma_name] = dict_name

        self.anal_namedict_r = {value: key for key, value in self.anal_namedict.items()}
    def df_clear(self):
        ''' 멤버변수 데이터프레임 초기화함수'''
        self.saved_df.drop(index=self.saved_df.index, columns=self.saved_df.columns, inplace=True)
        self.read_df_dayinfo.drop(index=self.read_df_dayinfo.index, columns=self.read_df_dayinfo.columns, inplace=True)
        self.read_df_criteria.drop(index=self.read_df_criteria.index, columns=self.read_df_criteria.columns, inplace=True)

    def module(self,code_update=False, day_info=False, daysub_info=False, compute_criteria=True, analysis=True, precent=2):
        self.mykrx.module(code_update, day_info, daysub_info)
        if compute_criteria == True:
            self.module_criteria_init()
        if analysis:
            self.analy_dict_update()
            self.make_scoreboard()
            self.module_analysis(percent=precent)
            self.score_while_file()
        #Scoring



    ''' 주식 계산 메소드 : Stock Scoring '''
        # A. Stock Analy 마지막 열에 계산된 값을 넣는다. --> 계산된 스코어의 마지막값 (오늘) 값은 dictionary 형태로 저장해둔다. {key = 점수 : value = [*Compony]}
        #       --> 지표 계산하면서 같이 열추가하는 것으로..
        # B. 점수 계산이 끝나면 오늘 기준으로 날짜시트를 만들어서 내림차순(높->낮) 으로 회사명 : 점수 이런식으로 엑셀파일 (Stock Score)로 추출한다.
    def scoring_each_compony(self, company : str):
        naming = "총점수"
        dic_list = self.saved_df.columns.tolist()
        self.saved_df[naming] = 0
        # 각 행마다 데이터를 적용하는 방법은 : apply 에 lambda 를 써서 해결하는 방법도 존재한다.

        idx_score = len(self.saved_df.columns) - 1
        for index in self.saved_df.index:
            for i, element in enumerate(self.saved_df.columns):
                try:
                    scores = self.anal_scoreboard[element]
                    value = scores[self.anal_score[self.saved_df.iloc[index, i]]]
                    if value == "EMPTY":
                        value = 0
                    self.saved_df.iloc[index, idx_score] += value
                except KeyError:
                    continue

        today_each_score = self.saved_df.tail(1)[naming].values[0]
        if today_each_score not in self.today_score:
            self.today_score[today_each_score] = [company]
        else:
            self.today_score[today_each_score].append(company)
        print("Final Score : " + str(today_each_score))

    def score_while_file(self):

        score_path = os.path.join(self.mykrx.subfile_path, self.analy_path["점수"])
        score_list = []
        sorted_keys = sorted(self.today_score.keys())
        for k in sorted_keys:
            for item in self.today_score[k]:
                score_list.append([item, k])
        score_df = pd.DataFrame(score_list, columns=["회사명", "총점수"])
        sheets_name = self.mykrx.today
        writeExcelFromDf(df=score_df, path=score_path, sname=sheets_name, index=True)

    def make_scoreboard(self):
        ''' 주식 점수지표 엑셀파일을 제작하는 메소드.'''
        score_file = os.path.join(self.mykrx.subfile_path, self.analy_path["점수"])
        self.make_default_scoreboard()
        sheet_name = "점수기준표"
        anal_score_df = pd.DataFrame(self.anal_scoreboard, index=self.anal_score.keys())
        if not os.path.isfile(score_file):
            anal_score_df.to_excel(excel_writer=score_file, sheet_name=sheet_name, index=True, header=True, engine="openpyxl")
            delete_empty_sheet(path=score_file)
        else:
            delete_empty_sheet(path=score_file)
            # Empty이외의 공간 값 비교
            tmp_dict = readExcelToDf(path=score_file, sname=sheet_name).to_dict()
            tmp_dict.pop("Unnamed: 0")
            for item, _ in self.anal_namedict_r.items():
                tmp_dict[item] = list(tmp_dict[item].values())

            changed = 0
            for item, _ in self.anal_namedict_r.items():
                for k,i in self.anal_score.items():
                    if self.anal_scoreboard[item][i] != tmp_dict[item][i]:
                        if (isinstance(self.anal_scoreboard[item][i], int) and isinstance(tmp_dict[item][i], str)) or\
                            (isinstance(self.anal_scoreboard[item][i], str) and isinstance(tmp_dict[item][i], int)):
                            print("[Score Strange : " + item + ":" + str(k) + "] " + str(tmp_dict[item][i]) + "-->" + str(self.anal_scoreboard[item][i]))
                        else:
                            print("[Score Change : " + item + ":" + str(k) + "] " + str(self.anal_scoreboard[item][i]) + "-->" + str(tmp_dict[item][i]))
                            changed = 1
            self.anal_scoreboard = tmp_dict
            if changed == True:
                print("[Scoring Changing Applying...]")
                anal_score_df = pd.DataFrame(self.anal_scoreboard, index=self.anal_score.keys())
                anal_score_df.to_excel(excel_writer=score_file, sheet_name=sheet_name, index=True, header=True, engine="openpyxl")
    def make_default_scoreboard(self):
        ''' 기본 점수표를 만드는 메소드'''
        self.analy_dict_update()
        #딕셔너리화 해서 재도전!

        score_list = list(self.anal_score.keys())
        for key, item in self.anal_namedict.items():
            self.anal_scoreboard[item] = ["EMPTY" for _ in range(len(self.anal_score))]
            '''
            tmp_dict = {i: "EMPTY" for _, i in self.anal_score.items()}
            self.anal_scoreboard[item] = tmp_dict
            '''

        self.anal_scoreboard[self.anal_namedict["SMA60_check"]][self.anal_score["up"]] = 1000
        self.anal_scoreboard[self.anal_namedict["SMA60_check"]][self.anal_score["down"]] = -1000

        self.anal_scoreboard[self.anal_namedict["전기_nearess_check(후행)"]][self.anal_score["O"]] = 1000
        self.anal_scoreboard[self.anal_namedict["전기_nearess_check(후행)"]][self.anal_score["X"]] = -1000

        self.anal_scoreboard[self.anal_namedict["전기_nearess_check"]][self.anal_score["O"]] = 100
        self.anal_scoreboard[self.anal_namedict["전기_nearess_check"]][self.anal_score["X"]] = -100

        for x in self.sma_window:
            name_s = "SMA" + str(x) + "_cross_check"
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["O"]] = 100
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["X"]] = -100

        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["up_near"]] = 500
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["up_cross"]] = 1000
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["up"]] = 700
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["down_near"]] = -1000
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["X"]] = -1000
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["down_cross"]] = -900
        self.anal_scoreboard[self.anal_namedict["MACD_check"]][self.anal_score["down"]] = -900

        for x in self.high_crit:
            name_s = str(x) + "_highest_check"
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["up_near"]] = 100
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["up_cross"]] = 70
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["up"]] = 50
            self.anal_scoreboard[self.anal_namedict[name_s]][self.anal_score["X"]] = -100

        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["up_near"]] = 50 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["up_cross"]] = 100 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["up"]] = 70 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["down_near"]] = -100 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["down_cross"]] = -90 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["down"]] = -90 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_line_cross_check"]][self.anal_score["X"]] = -100 * 2

        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["up_near"]] = 50 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["up_cross"]] = 100 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["up"]] = 70 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["down_near"]] = -100 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["down_cross"]] = -90 * 2
        self.anal_scoreboard[self.anal_namedict["후행스팬_bong_cross_check"]][self.anal_score["down"]] = -90 * 2

        self.anal_scoreboard[self.anal_namedict["스팬꼬리_check"]][self.anal_score["O"]] = 100
        self.anal_scoreboard[self.anal_namedict["스팬꼬리_check"]][self.anal_score["X"]] = -100

        self.anal_scoreboard[self.anal_namedict["스팬위치_check"]][self.anal_score["up"]] = 100
        self.anal_scoreboard[self.anal_namedict["스팬위치_check"]][self.anal_score["mid"]] = 0
        self.anal_scoreboard[self.anal_namedict["스팬위치_check"]][self.anal_score["down"]] = -100

        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["up_near"]] = 100
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["up_cross"]] = 80
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["up"]] = 50
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["down_near"]] = -50
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["down_cross"]] = -80
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["down"]] = -100
        self.anal_scoreboard[self.anal_namedict["전_cross_기"]][self.anal_score["X"]] = -100

        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["up_near"]] = 100
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["up_cross"]] = 70
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["up"]] = 50
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["down_near"]] = -50
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["down_cross"]] = -70
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["down"]] = -100
        self.anal_scoreboard[self.anal_namedict["봉_cross_전기"]][self.anal_score["X"]] = -100





    ''' 지표 분석 메소드'''
    def module_analysis(self, percent):
        self.df_clear()
        path_criteria = os.path.join(self.mykrx.subfile_path, self.analy_path["지표"])
        path_analy = os.path.join(self.mykrx.subfile_path, self.analy_path["분석"])
        path_dayinfo = os.path.join(self.mykrx.subfile_path, self.data_path["일봉"])
        if not os.path.isfile(path_criteria): #Compute 시작.
            self.module_criteria_init()

        print("Start Analysis")
        for compony, _ in self.mykrx.thema_total_dict.items():  #나중에 이 부분을 건들면 다른 딕셔너리에 대해서도 수행가능.
            print("[" + compony + " 지표 분석 중 ...]")
            self.read_df_dayinfo = readExcelToDf(path=path_dayinfo, sname=compony)
            self.read_df_criteria = readExcelToDf(path=path_criteria, sname=compony)
            print(self.read_df_criteria)
            self.saved_df["날짜"] = self.read_df_criteria["날짜"]

            #0. 60일 이동평균선의 추이
            self.sma60_direction(df_crit=self.read_df_criteria)
            #0-2. 전환선_기준선_가까움
            self.near_line_check(df_crit=self.read_df_criteria, percent=percent)
            #2. MACD 시그널 체킹 (하한~상한 제한걸어둠)
            self.check_macd(self.read_df_criteria)
            #4. 스팬친구들
            self.cross_backspan_line(df_crit=self.read_df_criteria, percent=percent)
            self.cross_backspan(df_day=self.read_df_dayinfo, df_crit=self.read_df_criteria)
            self.check_spantail(df_crit=self.read_df_criteria)
            self.check_span_position(df_day=self.read_df_dayinfo, df_crit=self.read_df_criteria)
            #5. 전환선 >= 기준선
            self.span_line_cross(df_crit=self.read_df_criteria, percent=percent)

            #6. 봉 >= 기준/전환선
            self.bong_cross_line(df_day=self.read_df_dayinfo, df_crit=self.read_df_criteria, percent=percent)
            #1. XX일 이동선을 통과했는가?
            self.cross_moving_line(df_day=self.read_df_dayinfo, df_crit=self.read_df_criteria)
            #3. *일 최고가 갱신여부?
            self.cross_highest_price(df_day=self.read_df_dayinfo, df_crit=self.read_df_criteria, percent=percent)
            #7. 점수화..
            self.scoring_each_compony(company=compony)
            print(self.saved_df.tail(5))
            #데이터 갱신
            add_excl_column(path=path_analy, sheets=compony, df=self.saved_df)
            self.df_clear()
    # 0. 60일 이동평균선의 추이 (up or Down)
    def sma60_direction(self, df_crit: pd.DataFrame, compare=[1,3,5,9,26]):
        naming = self.anal_namedict["SMA60_check"]
        self.saved_df[naming] = "down"
        mask = True
        for c in compare:
            mask &= (df_crit["SMA60"] > df_crit["SMA60"].shift(c))
        self.saved_df.loc[mask, naming] = "up"
    #0-1. 전환/기준선이 붙어있는가..
    def near_line_check(self, df_crit: pd.DataFrame, percent=2):
        naming2 = self.anal_namedict["전기_nearess_check"]
        naming = self.anal_namedict["전기_nearess_check(후행)"]
        self.saved_df[naming] = "X"
        mask = (df_crit[["전환선", "기준선"]].max(axis=1) * (1-0.01 * percent) < df_crit[["전환선", "기준선"]].min(axis=1))
        self.saved_df.loc[mask, naming] = "O"
        self.saved_df[naming2] = self.saved_df[naming].shift(25)
    #1. 평균이동선을 통과했는가?
    def cross_moving_line(self, df_day: pd.DataFrame, df_crit: pd.DataFrame):
        for sma in self.sma_window:
            sma_name = "SMA" + str(sma)
            d_key = sma_name + "_cross_check"
            naming = self.anal_namedict[d_key]
            self.saved_df[naming] = "X"
            mask = (df_day["종가"] >= df_crit[sma_name])
            self.saved_df.loc[mask, naming] = "O"
    #2. MACD 추이 (Up,Down_Cross)
    def check_macd(self, df_crit: pd.DataFrame(), low=1000):
        naming = self.anal_namedict["MACD_check"]
        #MACD 교차점을 상승하는지점 (매수지점 추천)
        self.saved_df[naming] = "X"
        mask = ((-low <= df_crit["MACD_Histogram"]))    #MACD 교차점이 -1000 ~ 1000 사이인가?
        mask &= (df_crit["MACD_Histogram"] > df_crit["MACD_Histogram"].shift(1))            #MACD 교차점이 5xPecrent 이하인가?
        mask &= (df_crit["MACD"] > 0)                                                       #MACD 값이 0보다 큰가?
        self.saved_df.loc[mask & (df_crit["MACD_Histogram"] > 0), naming] = "up"
        self.saved_df.loc[mask & (df_crit["MACD_Histogram"] < 0), naming] = "up_near"
        self.saved_df.loc[mask & ((df_crit["MACD_Histogram"].shift(1) < 0) & (df_crit["MACD_Histogram"] > 0)), naming] = "up_cross"

        mask2 = ((low > df_crit["MACD_Histogram"]))    #MACD 교차점이 5xPecrent 이하인가?
        mask2 &= (df_crit["MACD_Histogram"] < df_crit["MACD_Histogram"].shift(1))            #MACD 교차점의 하강중인가?
        self.saved_df.loc[mask2 & (df_crit["MACD_Histogram"] < 0), naming] = "down"  #MACD 값이 0보다 큰가?
        self.saved_df.loc[mask2 & (df_crit["MACD_Histogram"] > 0), naming] = "down_near"
        self.saved_df.loc[mask2 & ((df_crit["MACD_Histogram"].shift(1) > 0) & (df_crit["MACD_Histogram"] < 0)), naming] = "down_cross"

        #MACD 교차점을 하강하는지점 (매도지점 추천)
    #3. &일 최고가를 갱신했는가?
    def cross_highest_price(self, df_day: pd.DataFrame, df_crit: pd.DataFrame, percent=2):
        ''' 최고가 기준을 Cross 했는가? (근접했으면 near, 같거나 띄어넘었으면 up, 크로스했으면 up_cross 아니면 X)
        :param df_day: 일봉 데이터프레임
        :param df_crit: 기준표 데이터프레임
        :param percent: 퍼센트 기준
        :return:
        '''
        for hi in self.high_crit:
            d_key = str(hi) + "_highest_check"
            df_key = str(hi) + "일_최고가"
            naming = self.anal_namedict[d_key]
            self.saved_df[naming] = "X"
            mask = (df_day["종가"] >= df_crit[df_key] * (1 - 0.01*percent))
            self.saved_df.loc[mask & (df_day["종가"] < df_crit[df_key]), naming] = "up_near"
            self.saved_df.loc[mask & (df_day["종가"] >= df_crit[df_key]), naming] = "up"
            self.saved_df.loc[mask & (df_day["종가"].shift(1) < df_crit[df_key].shift(1)) &\
                              (df_day["종가"] >= df_crit[df_key]), naming] = "up_cross"

    #4-01. 후행스팬이 기준선,전환선을 Cross할 때
    def cross_backspan_line(self, df_crit: pd.DataFrame, percent=2):
        naming = self.anal_namedict["후행스팬_line_cross_check"]
        naming2 = self.anal_namedict["전기_nearess_check(후행)"]
        self.saved_df[naming] = "X"
        mask = (df_crit["후행스팬"] >= df_crit[["전환선", "기준선"]].min(axis=1) * (1 - 0.01 * percent))
        mask &= (self.saved_df[naming2] == "O")
        self.saved_df.loc[
            mask & (df_crit["후행스팬"] < df_crit[["전환선", "기준선"]].min(axis=1)),
            naming
        ] = "up_near"
        self.saved_df.loc[
            mask & (df_crit["후행스팬"] >= df_crit[["전환선", "기준선"]].min(axis=1)),
            naming
        ] = "up"
        self.saved_df.loc[
            mask & (df_crit["후행스팬"] >= df_crit[["전환선", "기준선"]].min(axis=1)) & (
                    df_crit["후행스팬"].shift(1) < df_crit[["전환선", "기준선"]].min(axis=1).shift(1)),
            naming
        ] = "up_cross"

        mask2 = (df_crit["후행스팬"] < df_crit[["전환선", "기준선"]].max(axis=1) * (1 - 0.01 * percent))
        self.saved_df.loc[
            mask2 & (df_crit["후행스팬"] > df_crit[["전환선", "기준선"]].max(axis=1)),
            naming
        ] = "down_near"
        self.saved_df.loc[
            mask2 & (df_crit["후행스팬"] < df_crit[["전환선", "기준선"]].max(axis=1)),
            naming
        ] = "down"
        self.saved_df.loc[
            mask2 & (df_crit["후행스팬"] <= df_crit[["전환선", "기준선"]].max(axis=1)) & (
                        df_crit["후행스팬"].shift(1) > df_crit[["전환선", "기준선"]].max(axis=1).shift(1)),
            naming
        ] = "down_cross"
        self.saved_df[naming] =  self.saved_df[naming].shift(25)
    #4-1. 26일전 주식 종가가 후행스팬을 뚫었나?
    def cross_backspan(self, df_day: pd.DataFrame, df_crit: pd.DataFrame, percent=2):

        check_naming = self.anal_namedict["후행스팬_bong_cross_check"]
        self.saved_df[check_naming] = "X"
        mask = (df_day["종가"] >= df_crit["후행스팬"] * (1-0.01 * percent))
        self.saved_df.loc[
            mask & (df_day["종가"] < df_crit["후행스팬"]),
            check_naming
        ] = "up_near"
        self.saved_df.loc[
            mask & (df_day["종가"] >= df_crit["후행스팬"]),
            check_naming
        ] = "up"
        self.saved_df.loc[
            mask & (df_day["종가"] > df_crit["후행스팬"]) & (df_day["종가"].shift(1) < df_crit["후행스팬"].shift(1)),
            check_naming
        ] = "up_cross"

        mask2 = (df_day["종가"] <= df_crit["후행스팬"] * (1 - 0.01 * percent))
        self.saved_df.loc[
            mask2 & (df_day["종가"] > df_crit["후행스팬"]),
            check_naming
        ] = "down_near"
        self.saved_df.loc[
            mask2 & (df_day["종가"] <= df_crit["후행스팬"]),
            check_naming
        ] = "down"
        self.saved_df.loc[
            mask2 & (df_day["종가"] < df_crit["후행스팬"]) & (df_day["종가"].shift(1) > df_crit["후행스팬"].shift(1)),
            check_naming
        ] = "down_cross"
        self.saved_df[check_naming] = self.saved_df[check_naming].shift(25)

    #4-2. 현 주가의 스팬 꼬리(3일전부터 모두 상승?) 가 양의 방향인가? + 스팬이 양수인가?
    def check_spantail(self, df_crit: pd.DataFrame):
        naming = self.anal_namedict["스팬꼬리_check"]
        self.saved_df[naming] = "X"
        mask = (df_crit["선행스팬1_미래"] > df_crit["선행스팬2_미래"])
        mask &= (df_crit["선행스팬1_미래"] > df_crit["선행스팬1_미래"].shift(1))
        mask &= (df_crit["선행스팬1_미래"].shift(1) > df_crit["선행스팬1_미래"].shift(2))
        self.saved_df.loc[mask, naming] = "O"
    #4-3. 현 주가의 스팬 위치 (up, mid, bot)
    def check_span_position(self, df_day: pd.DataFrame, df_crit: pd.DataFrame):
        naming = self.anal_namedict["스팬위치_check"]
        self.saved_df[naming] = "mid"

        self.saved_df.loc[(df_day["종가"] <= df_crit[["선행스팬1","선행스팬2"]].min(axis=1)), naming] = "down"
        self.saved_df.loc[(df_day["종가"] >= df_crit[["선행스팬1","선행스팬2"]].max(axis=1)), naming] = "up"

    #5. 전환선 >= 기준선 (통과전, 통과, 통과 후)
    def span_line_cross(self, df_crit: pd.DataFrame, percent=2):
        naming = self.anal_namedict["전_cross_기"]
        self.saved_df[naming] = "X"
        mask = (df_crit["전환선"] >= df_crit["기준선"] * (1 - 0.01 * percent))
        self.saved_df.loc[mask & (df_crit["전환선"] < df_crit["기준선"]), naming] = "up_near"
        self.saved_df.loc[mask & (df_crit["전환선"] >= df_crit["기준선"]), naming] = "up"
        self.saved_df.loc[mask & (df_crit["전환선"].shift(26) < df_crit["기준선"].shift(26)) & \
                          (df_crit["전환선"].shift(25) >= df_crit["기준선"].shift(25)), naming] = "up_cross"

        mask2 = (df_crit["전환선"] < df_crit["기준선"] * (1 - 0.01 * percent))
        self.saved_df.loc[mask2 & (df_crit["전환선"] > df_crit["기준선"]), naming] = "down_near"
        self.saved_df.loc[mask2 & (df_crit["전환선"] <= df_crit["기준선"]), naming] = "down"
        self.saved_df.loc[mask2 & (df_crit["전환선"].shift(26) > df_crit["기준선"].shift(26)) & \
                          (df_crit["전환선"].shift(25) <= df_crit["기준선"].shift(25)), naming] = "down_cross"

    #6. 봉이 기준선과 전환선을 뚫고 갈려고 하는가? (+ 기준선+전환선이 붙어있어야함)
    def bong_cross_line(self, df_day: pd.DataFrame, df_crit: pd.DataFrame, percent=2):
        naming = self.anal_namedict["봉_cross_전기"]
        naming2 = self.anal_namedict["전기_nearess_check"]
        self.saved_df[naming] = "X"
        mask = (df_day["종가"] >= df_crit[["전환선", "기준선"]].min(axis=1) * (1 - 0.01 * percent))
        mask &= (self.saved_df[naming2] == "O")
        self.saved_df.loc[
            mask & (df_day["종가"] < df_crit[["전환선", "기준선"]].min(axis=1)),
            naming
        ] = "up_near"
        self.saved_df.loc[
            mask & (df_day["종가"] >= df_crit[["전환선", "기준선"]].min(axis=1)),
            naming
        ] = "up"
        self.saved_df.loc[
            mask & (df_day["종가"] >= df_crit[["전환선", "기준선"]].min(axis=1)) & (
                    df_day["종가"].shift(1) < df_crit[["전환선", "기준선"]].min(axis=1).shift(1)),
            naming
        ] = "up_cross"

        mask2 = (df_day["종가"] < df_crit[["전환선", "기준선"]].max(axis=1) * (1 - 0.01 * percent))
        self.saved_df.loc[
            mask2 & (df_day["종가"] > df_crit[["전환선", "기준선"]].max(axis=1)),
            naming
        ] = "down_near"
        self.saved_df.loc[
            mask2 & (df_day["종가"] < df_crit[["전환선", "기준선"]].max(axis=1)),
            naming
        ] = "down"
        self.saved_df.loc[
            mask2 & (df_day["종가"] <= df_crit[["전환선", "기준선"]].max(axis=1)) & (
                        df_day["종가"].shift(1) > df_crit[["전환선", "기준선"]].max(axis=1).shift(1)),
            naming
        ] = "down_cross"


    ''' 지표 계산 메소드 ...'''
    def module_criteria_init(self):
        ''' 초기 day_info.xlsx 로부터 MACD, 60평균선 등을 계산하는데 사용하는 함수 (전데이터 수정)
        :param hgih_crit: 최고가 날 기준 (기본값 240일)
        :return:
        '''
        path_data = os.path.join(self.mykrx.subfile_path, self.data_path["일봉"])
        path_analy = os.path.join(self.mykrx.subfile_path, self.analy_path["지표"])
        for company, _ in self.mykrx.thema_total_dict.items():
            print("["+company+" 지표 계산중...]")
            df = readExcelToDf(path=path_data, sname=company)
            self.saved_df["날짜"] = df["날짜"]

            #1. 주가이동평균 구함.
            print("{주가이동평균(Moving Average) 계산 중 ...}")
            self.movingAverage(cal_df=df)

            #2. MACD 구함.
            print("{MACD(Moving Average Convergence Divergence) 계산 중 ...}")
            self.macd(cal_df=df)

            #3. 일목기준표 구함.
            print("{일목균형표(Ichimoku Kinkoyo) 계산 중 ...}")
            self.ichimoku(cal_df=df)

            #4. X일 중 최고가를 구함.
            print("{"+ str(self.high_crit) +"일 최고가(highest price) 계산 중 ...}")
            self.highest_price(cal_df=df)
            print(self.saved_df.tail(5))
            add_excl_column(path=path_analy, df=self.saved_df, sheets=company)
            self.df_clear()
    #이동평균선 구하는 메소드
    def movingAverage(self, cal_df: pd.DataFrame):
        ''' self.saved_df 에 저장된 데이터로 기반으로 이동평균(moving Average)를 계산 '''
        for w in self.sma_window:
            col_name = "SMA"+str(w)
            self.saved_df[col_name] = ta.SMA(cal_df["종가"], timeperiod=w)
        return self.saved_df
    #MACD 구하는 메소드
    def macd(self, cal_df: pd.DataFrame):
        ema12 = ta.EMA(cal_df["종가"], timeperiod=self.ema_window[1])
        ema26 = ta.EMA(cal_df["종가"], timeperiod=self.ema_window[2])
        self.saved_df["MACD"] = ema12 - ema26
        self.saved_df["MACD_Signal"] = ta.EMA(self.saved_df["MACD"], timeperiod=self.ema_window[0])
        self.saved_df["MACD_Histogram"] = self.saved_df["MACD"] - self.saved_df["MACD_Signal"]
        return self.saved_df
    #일목스님 기준표
    def ichimoku(self, cal_df: pd.DataFrame):
        '''
        고가	저가
        전환선: 9일간의 최고가 + 최소가 의 평균
        기준선: 26일간의 최고 + 최소 의 평균
        선행스팬1: 기준선(Kijun-sen)을 26일 전으로 이동시킵니다.
        선행스팬2: 최근 52일의 고가(High)와 저가(Low)를 더한 후, 52로 나눈 값을 26일 전으로 이동시킵니다.
        후행스팬: 현재 주가를 26일 전으로 이동시킵니다.
        '''
        self.saved_df["전환선"] = (cal_df["고가"].rolling(window=9).max() + cal_df["저가"].rolling(window=9).min()) / 2
        self.saved_df["기준선"] = (cal_df["고가"].rolling(window=26).max() + cal_df["저가"].rolling(window=26).min()) / 2
        self.saved_df["선행스팬1"] = ((self.saved_df["기준선"] + self.saved_df["전환선"])/2).shift(26)
        self.saved_df["선행스팬2"] = ((cal_df["고가"].rolling(window=52).max() + cal_df["저가"].rolling(window=52).min()) / 2).shift(26)
        self.saved_df["후행스팬"] = cal_df["종가"].shift(-25)
        self.saved_df["선행스팬1_미래"] = cal_df["종가"] = ((self.saved_df["기준선"] + self.saved_df["전환선"])/2)
        self.saved_df["선행스팬2_미래"] = ((cal_df["고가"].rolling(window=52).max() + cal_df["저가"].rolling(window=52).min()) / 2)
    #X일 가장 고가? --> 9 26
    def highest_price(self, cal_df: pd.DataFrame):
        ''' high_crit 일 종가 중 최고가?
        :param cal_df:
        :param high_crit:
        :return:
        '''
        for high in self.high_crit:
            naming = str(high) + "일_최고가"
            self.saved_df[naming] = cal_df["종가"].rolling(window=high).max()

if __name__ == "__main__":
    analy = StockAnaly()
    analy.module(day_info=False, compute_criteria=True, analysis=True)

