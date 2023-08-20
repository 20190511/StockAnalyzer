# DataFrame 엑셀 보조 모듈 파일
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

def e_delete_sheet(filename: str, sname: str):
    if not os.path.isfile(filename):
        return
    wb = openpyxl.load_workbook(filename=filename)
    if sname in wb.sheetnames:
        del wb[sname]
    wb.save(filename)
    return
def e_read(filename: str, sname: str, last=False,
           min_row=-1, max_row=-1, min_col=-1, max_col=-1):
    ''' Excell Read 보조함수.
    :param filename: 파일명
    :param sname: 시트명 (없으면 "")
    :param last:  마지막 행 값을 가져오는가?
    :param min_row: 최소범위 행
    :param max_row: 최대범위 행
    :param min_col: 최소범위 열
    :param max_col: 최대범위 행
    :return: pd.DataFrame(가져온 데이터)
    '''
    e_df = pd.DataFrame
    if not os.path.isfile(filename):
        return e_df

    wb = openpyxl.load_workbook(filename=filename)

    global  ws
    if sname == "":
        ws = wb.active()
    else:
        if not sname in wb.sheetnames:
            return e_df
        ws = wb[sname]

    #openpyxl : sheets에서 iter_rows를 하면 --> 행 단위로 시트를 볼 수 있음.
        #ex) iter_rows(min_col = 1, max_col = sheet.
    if min_row == -1:
        min_row = 0
    if max_row >= ws.max_row or max_row == -1:
        max_row = ws.max_row
    if min_col == -1:
        min_col = 0
    if max_col >= ws.max_column or max_col == -1:
        max_col = ws.max_column
    # 마지막 값만 가져옴.
    if last == True:
        min_row = ws.max_row
        max_row = ws.max_row
    df = pd.DataFrame(ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=True))

    if min_row == 0:
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)
        df.set_index(df.columns[0], inplace=True)

    return df
def e_write(df: pd.DataFrame, filename: str, sname: str, last=True, last_col=False, index=False, header=False):
    ''' 엑셀 Write 모듈함수
    :param df:          DataFrame.
    :param filename:    filename
    :param sname:       sheet name
    :param last:        마지막 행에 데이터를 쓰는가 여부?
    :param last_col:    엑셀 열 추가 옵션 : df에 추가할 열을 넣어주면 알아서 columns 분해해서 처리함.
    :param index:       index 옵션
    :param header:      엑셀 맨 위에 header 여부
    :return:
    '''
    global ws

    if not os.path.isfile(filename):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(filename=filename)

    if last == False:
        if sname in wb.sheetnames:
            del wb[sname]
        header = True

    if not sname in wb.sheetnames:
        wb.create_sheet(sname)
    ws = wb[sname]

    df = df.reset_index()
    if last_col == True:
        header = True
        index_num = []
        print(df.columns)
        max_col_num = ws.max_column
        if max_col_num != 1:
            max_col_num += 1
        for i in range(max_col_num, max_col_num+1+len(df.columns)):
            index_num.append(i)
            ws.insert_cols(i)

        count = 1
        for row in dataframe_to_rows(df=df, index=index, header=header):
            for i, row_col in enumerate(row):
                ws.cell(row=count, column=index_num[i], value=row_col)
            count += 1

    else:
        for row in dataframe_to_rows(df=df, index=index, header=header):
            ws.append(row)

    wb.save(filename=filename)


def e_wb(filename: str):
    if not os.path.isfile(filename):
        return openpyxl.Workbook()
    else:
        return openpyxl.load_workbook(filename=filename)
class SavedExcel:
    ''' 파일 저장을 늦게시행해주는 메소드'''
    def __init__(self, filename: str):
        self.filename = filename
        self.saved_wb = e_wb(filename)
        self.saved_ws = self.saved_wb.active

    def write(self, df: pd.DataFrame, sname: str, last=True, last_col=False, index=False, header=False):
        if last == False:
            if sname in self.saved_wb.sheetnames:
                del self.saved_wb[sname]
            header = True

        if not sname in self.saved_wb.sheetnames:
            self.saved_wb.create_sheet(sname)

        self.saved_ws = self.saved_wb[sname]
        if "Sheet" in self.saved_wb.sheetnames:
            del self.saved_wb["Sheet"]

        df = df.reset_index()
        if last_col == True:
            header = True
            index_num = []
            max_col_num = self.saved_ws.max_column
            if max_col_num != 1:
                max_col_num += 1
            for i in range(max_col_num, max_col_num + len(df.columns)):
                if i != 1:
                    self.saved_ws.insert_cols(i)
                index_num.append(i)

            count = 1
            for row in dataframe_to_rows(df=df, index=index, header=header):
                for i, row_col in enumerate(row):
                    self.saved_ws.cell(row=count, column=index_num[i], value=row_col)
                count += 1

        else:
            for row in dataframe_to_rows(df=df, index=index, header=header):
                self.saved_ws.append(row)

    def saved(self):
        self.saved_wb.save(filename=self.filename)
    def open(self):
        self.saved_wb = openpyxl.load_workbook(filename=self.filename)


if __name__ == "__main__":
    path = os.path.join(os.getcwd(), "서브파일", "dayinfo.xlsx")
    path2 = os.path.join(os.getcwd(), "서브파일", "test2.xlsx")

    '''
    lista = [
        ["준형", 20, "컴공"],
        ["주협", 20, "컴공"],
        ["윤호", 20, "신소재"],
    ]
    df = pd.DataFrame(data=lista, columns=["이름","나이","전공"])
    df.set_index(df.columns[0], inplace=True)
    s = SavedExcel(filename=path2)
    s.write(df=df, last_col=True, sname="실험", header=True, index=False)
    s.write(df=df, last_col=True, sname="실험", header=True, index=False)
    s.write(df=df, last_col=True, sname="실험", header=True, index=False)
    s.saved()
    '''